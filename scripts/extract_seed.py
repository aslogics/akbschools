"""Extract AKB admission workbook -> data/students.seed.json and data/seed.js

Fee heads (matching the Chairman Dashboard), with their business owner:
  term (Terms Fees)            AKB School of Excellence   cols T/U/V   (20/21/22)
  supplies (School Supplies)   AKB & Co                   cols W/X/Y   (23/24/25)
  app_fees (App Fees Paid)     AKB School of Excellence   col  Z      (26, paid only)
  uniform (Uniform & Acces)    AKB & Co                   cols AA/AB/AC(27/28/29)
  transport (Transport Fees)   Falcon Trading & Transport cols AD/AE/AF(30/31/32)
  extra_curricular             AKB School of Excellence   cols AG/AH/AI(33/34/35)
  evening_sports               AKB School of Excellence   cols AS/AT/AU(45/46/47)
Marks: English AM(39) Maths AN(40) Science AO(41); Sports activity AP(42)

Usage:  python3 scripts/extract_seed.py /path/to/workbook.xlsx
"""
import openpyxl, json, math, datetime, sys, os

SRC = sys.argv[1] if len(sys.argv) > 1 else \
    '/root/.claude/uploads/9fd7b7df-9bc7-50a6-8565-5c8fdd6af6d5/df4d262f-AKB_ADMISSION_2026_TO_2027NEW.xlsx'
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['DATA']


def num(v):
    if v is None or v == '':
        return 0
    if isinstance(v, (int, float)):
        return 0 if (isinstance(v, float) and math.isnan(v)) else round(v, 2)
    try:
        return round(float(str(v).replace(',', '').strip()), 2)
    except Exception:
        return 0


def txt(v):
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# key, label, total_col, paid_col, balance_col (None => single paid value)
HEADS = [
    ('term', 'Terms Fees', 20, 21, 22),
    ('supplies', 'School Supplies', 23, 24, 25),
    ('app_fees', 'App Fees Paid', 26, 26, None),
    ('uniform', 'Uniform & Accessories', 27, 28, 29),
    ('transport', 'Transport Fees', 30, 31, 32),
    ('extra_curricular', 'Extra Curricular Fees', 33, 34, 35),
    ('evening_sports', 'Evening Sports', 45, 46, 47),
]

students = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 3).value
    sid = txt(ws.cell(r, 4).value)
    if not name or not str(name).strip() or not sid:
        continue
    heads = {}
    for key, label, tc, pc, bc in HEADS:
        total = num(ws.cell(r, tc).value)
        paid = num(ws.cell(r, pc).value)
        if bc is None:            # single paid value (App Fees)
            total = paid
        elif total == 0 and paid:  # derive total if missing
            total = paid
        heads[key] = {'label': label, 'total': total, 'paid': paid,
                      'balance': round(total - paid, 2)}
    marks = {
        'english': txt(ws.cell(r, 39).value),
        'maths': txt(ws.cell(r, 40).value),
        'science': txt(ws.cell(r, 41).value),
    }
    students.append({
        'id': sid,
        'name': txt(ws.cell(r, 3).value),
        'grade': txt(ws.cell(r, 5).value),
        'classTeacher': txt(ws.cell(r, 6).value),
        'gender': txt(ws.cell(r, 7).value),
        'dob': txt(ws.cell(r, 8).value),
        'age': txt(ws.cell(r, 9).value),
        'prevSchool': txt(ws.cell(r, 10).value),
        'father': txt(ws.cell(r, 11).value),
        'mother': txt(ws.cell(r, 12).value),
        'location': txt(ws.cell(r, 13).value),
        'dropLocation': txt(ws.cell(r, 14).value),
        'transportType': txt(ws.cell(r, 15).value),
        'vehicle': txt(ws.cell(r, 16).value),
        'contact': txt(ws.cell(r, 17).value),
        'religion': txt(ws.cell(r, 18).value),
        'discount': num(ws.cell(r, 19).value),
        'admission': txt(ws.cell(r, 43).value),
        'sportsActivity': txt(ws.cell(r, 42).value),
        'marks': marks,
        'fees': heads,
    })

payload = {'school': 'AKB School of Excellence', 'year': '2026-2027',
           'generated': 'seed', 'students': students}

json_path = os.path.join(ROOT, 'data', 'students.seed.json')
js_path = os.path.join(ROOT, 'data', 'seed.js')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(payload, f, indent=1, ensure_ascii=False)
with open(js_path, 'w', encoding='utf-8') as f:
    f.write('// Auto-generated from AKB admission workbook. Regenerate: '
            'python3 scripts/extract_seed.py <workbook.xlsx>\n')
    f.write('window.__AKB_SEED__ = ')
    json.dump(payload, f, ensure_ascii=False)
    f.write(';\n')

tot = sum(sum(h['total'] for h in s['fees'].values()) for s in students)
paid = sum(sum(h['paid'] for h in s['fees'].values()) for s in students)
print('students:', len(students))
print('total fees:', int(tot), 'paid:', int(paid), 'balance:', int(tot - paid))
print('wrote', json_path, 'and', js_path)
